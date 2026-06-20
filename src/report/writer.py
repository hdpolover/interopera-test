# src/report/writer.py
"""Report writer — populates sample_docs/report_template.xlsx from computed figures only.

Loads the provided report_template.xlsx (Section + Metric pre-filled by the brief),
writes Value / Limit / Utilization / Status / Source into columns C–G, and applies
status-based row colouring (BREACH=red, AT LIMIT=amber, OK=green).

All cells sourced exclusively from the figures list — no narrative/LLM input.
Falls back to generating a new workbook if the template is not found.
"""
from __future__ import annotations

import os
from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill

from src.compute.registry import Figure

# openpyxl copies style objects into each cell on assignment, so these
# module-level instances are safely reused across multiple workbooks and cells.
_STATUS_FILL: dict[str, PatternFill] = {
    "BREACH":   PatternFill(fill_type="solid", fgColor="FF4444"),
    "AT LIMIT": PatternFill(fill_type="solid", fgColor="FFAA00"),
    "OK":       PatternFill(fill_type="solid", fgColor="00AA44"),
}
_WHITE_FONT = Font(color="FFFFFF")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="EEEEEE")
_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="center")

# Template lives in sample_docs/ relative to repo root.
# In Docker the repo is mounted at /app; locally it is the CWD.
_TEMPLATE_NAME = os.path.join("sample_docs", "report_template.xlsx")

# Ordered list of (section, metric_label, figure_id) matching report_template.xlsx row order.
# Inverse of reconciler._METRIC_TO_FIGURE_ID, using the canonical template metric names.
_TEMPLATE_ROWS: list[tuple[str, str, str]] = [
    ("Allocation",    "Singapore Government Securities",      "allocation_sgs"),
    ("Allocation",    "MAS Bills",                            "allocation_mas_bills"),
    ("Allocation",    "Investment Grade Corporate Bonds",     "allocation_ig_corp"),
    ("Allocation",    "High Yield Bonds",                     "allocation_high_yield"),
    ("Allocation",    "Foreign Currency Bonds (hedged)",      "allocation_fx_bonds"),
    ("Allocation",    "Structured Credit (ABS/MBS)",          "allocation_structured_credit"),
    ("Allocation",    "Cash & Cash Equivalents",              "allocation_cash"),
    ("Aggregate",     "Aggregate non-IG exposure",            "aggregate_non_ig_exposure"),
    ("Concentration", "Largest single corporate issuer",      "largest_single_corporate_issuer"),
    ("Concentration", "Largest GRE issuer",                   "largest_gre_issuer"),
    ("Liquidity",     "Liquid assets ratio",                  "liquid_assets_ratio"),
    ("Market risk",   "Portfolio modified duration",          "portfolio_duration"),
    ("Market risk",   "Portfolio DV01",                       "portfolio_dv01"),
]

_HEADERS: list[str] = [
    "Section",
    "Metric",
    "Value",
    "Limit",
    "Utilization",
    "Status",
    "Source (graph path → doc/page)",
]


def _build_source(fig: Figure) -> str:
    """Build the Source cell content from graph_path and citation fields only."""
    parts: list[str] = []
    if fig.graph_path:
        parts.append(fig.graph_path)
    citation = fig.citation or {}
    source_doc = citation.get("source_doc", "")
    page = citation.get("page", "")
    chunk_id = citation.get("chunk_id", "")
    citation_parts: list[str] = []
    if source_doc:
        page_str = f" p.{page}" if page != "" else ""
        citation_parts.append(f"{source_doc}{page_str}")
    if chunk_id:
        citation_parts.append(str(chunk_id))
    if citation_parts:
        parts.append(" ".join(citation_parts))
    return " | ".join(parts)


def _find_template() -> str | None:
    """Search for report_template.xlsx from CWD upward (handles Docker /app and local)."""
    candidates = [
        Path(_TEMPLATE_NAME),
        Path("/app") / _TEMPLATE_NAME,
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


def write_report(figures: list[Figure], output_path: str) -> None:
    """Populate report_template.xlsx with computed figures and save to output_path.

    Loads sample_docs/report_template.xlsx (Section + Metric pre-filled by the brief)
    and writes Value / Limit / Utilization / Status / Source into columns C–G for
    each of the 13 metric rows. Falls back to generating a new workbook if the
    template is not found.

    Formatting applied:
    - Header row: bold + light gray fill (#EEEEEE)
    - BREACH rows: red background (#FF4444) + white text
    - AT LIMIT rows: amber background (#FFAA00) + white text
    - OK rows: green background (#00AA44) + white text
    - Auto-fit column widths based on max content length

    All cells sourced exclusively from the figures list — no narrative/LLM input.
    """
    import openpyxl

    fig_map: dict[str, Figure] = {f.figure: f for f in figures}
    template_path = _find_template()

    if template_path:
        # Populate the provided template — Section + Metric already in cols A + B.
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Style the header row (row 1 already has content from template).
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        # Fill columns C–G (3–7) for each data row, in template row order.
        for row_idx, (_, _, fig_id) in enumerate(_TEMPLATE_ROWS, start=2):
            fig = fig_map.get(fig_id)
            status = None
            if fig is not None:
                ws.cell(row=row_idx, column=3).value = fig.value
                ws.cell(row=row_idx, column=4).value = fig.limit
                ws.cell(row=row_idx, column=5).value = fig.utilization
                ws.cell(row=row_idx, column=6).value = fig.status
                ws.cell(row=row_idx, column=7).value = _build_source(fig)
                status = fig.status

            if status in _STATUS_FILL:
                fill = _STATUS_FILL[status]
                for col in range(1, 8):
                    ws.cell(row=row_idx, column=col).fill = fill
                    ws.cell(row=row_idx, column=col).font = _WHITE_FONT

    else:
        # Fallback: generate workbook from scratch (template not found).
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        ws.append(_HEADERS)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        for section, metric, fig_id in _TEMPLATE_ROWS:
            fig = fig_map.get(fig_id)
            if fig is not None:
                row_data = [section, metric, fig.value, fig.limit,
                            fig.utilization, fig.status, _build_source(fig)]
                status = fig.status
            else:
                row_data = [section, metric, None, None, None, None, None]  # type: ignore[list-item]
                status = None
            ws.append(row_data)
            if status in _STATUS_FILL:
                for cell in ws[ws.max_row]:
                    cell.fill = _STATUS_FILL[status]
                    cell.font = _WHITE_FONT

    # Auto-fit column widths.
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[col_cells[0].column_letter].width = max_len + 4

    wb.save(output_path)
