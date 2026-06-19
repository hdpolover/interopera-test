"""Report writer tests."""
import os
import tempfile
import pytest
from src.compute.registry import Figure


@pytest.fixture
def sample_figures():
    return [
        Figure(figure="allocation_sgs", value="35.0%", utilization="58.3%", status="OK",
               limit="20–60%", graph_path="(Position:SGS-01,SGS-02)->AssetClass", citation={"chunk_id": "abc12345"}),
        Figure(figure="allocation_cash", value="4.0%", utilization="n/a", status="BREACH",
               limit="min 5%", graph_path="(Position:CASH-01)->AssetClass", citation={"chunk_id": "def67890"}),
    ]


def test_write_report_creates_xlsx(sample_figures):
    from src.report.writer import write_report
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_report(sample_figures, path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
    finally:
        os.unlink(path)


def test_write_report_has_13_data_rows():
    from src.report.writer import write_report
    from src.compute.registry import Figure
    figures_13 = [
        Figure(figure=f"fig_{i}", value=f"{i}.0%", utilization="n/a", status="OK",
               limit="0–100%", graph_path="", citation={})
        for i in range(13)
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_report(figures_13, path)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # First row is header; data starts at row 2
        data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(c for c in row)]
        assert len(data_rows) == 13
    finally:
        os.unlink(path)


def test_write_report_values_come_from_figures(sample_figures):
    from src.report.writer import write_report
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_report(sample_figures, path)
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        values_in_sheet = set()
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    values_in_sheet.add(str(cell))
        assert "35.0%" in values_in_sheet
        assert "4.0%" in values_in_sheet
        assert "BREACH" in values_in_sheet
    finally:
        os.unlink(path)


def test_write_report_signature_has_no_narrative_param():
    import inspect
    from src.report.writer import write_report
    sig = inspect.signature(write_report)
    assert "narrative" not in sig.parameters


def test_write_report_columns_match_template():
    """Headers must match report_template.xlsx exactly."""
    from src.report.writer import write_report
    import openpyxl
    figures = [
        Figure(figure="allocation_sgs", value="35.0%", utilization="58.3%", status="OK",
               limit="20–60%", graph_path="(Position:SGS-01)->AssetClass",
               citation={"source_doc": "guidelines.pdf", "page": 3, "chunk_id": "x1"}),
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_report(figures, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        expected_headers = [
            "Section", "Metric", "Value", "Limit", "Utilization", "Status",
            "Source (graph path → doc/page)",
        ]
        assert headers == expected_headers, f"Headers mismatch: {headers}"
    finally:
        os.unlink(path)


def test_write_report_source_column_contains_graph_path_and_citation():
    """Source column must include graph_path and citation info."""
    from src.report.writer import write_report
    import openpyxl
    figures = [
        Figure(
            figure="allocation_sgs",
            value="35.0%",
            utilization="58.3%",
            status="OK",
            limit="20–60%",
            graph_path="(Position:SGS-01,SGS-02)->AssetClass",
            citation={"source_doc": "guidelines.pdf", "page": 3, "chunk_id": "abc12345"},
        ),
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_report(figures, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Find the row for allocation_sgs (Singapore Government Securities)
        source_val = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == "Singapore Government Securities":
                source_val = row[6]  # Source column is index 6
                break
        assert source_val is not None, "Row for SGS not found"
        assert "(Position:SGS-01,SGS-02)->AssetClass" in source_val
        assert "abc12345" in source_val
    finally:
        os.unlink(path)


def test_write_report_template_metric_rows_present():
    """All 13 template metric rows must be present in output."""
    from src.report.writer import write_report
    import openpyxl
    from src.compute.registry import FIGURE_REGISTRY
    # Build figures for all 13 spec IDs
    figures = [
        Figure(
            figure=spec.id,
            value="10.0%",
            utilization="50.0%",
            status="OK",
            limit=spec.limit_display,
            graph_path=f"graph/{spec.id}",
            citation={"source_doc": "doc.pdf", "page": 1, "chunk_id": f"cid_{spec.id}"},
        )
        for spec in FIGURE_REGISTRY
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        write_report(figures, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        metrics_in_sheet = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                metrics_in_sheet.append(row[1])
        expected_metrics = [
            "Singapore Government Securities",
            "MAS Bills",
            "Investment Grade Corporate Bonds",
            "High Yield Bonds",
            "Foreign Currency Bonds (hedged)",
            "Structured Credit (ABS/MBS)",
            "Cash & Cash Equivalents",
            "Aggregate non-IG exposure",
            "Largest single corporate issuer",
            "Largest GRE issuer",
            "Liquid assets ratio",
            "Portfolio modified duration",
            "Portfolio DV01",
        ]
        assert metrics_in_sheet == expected_metrics, f"Metrics mismatch:\n{metrics_in_sheet}\n!=\n{expected_metrics}"
    finally:
        os.unlink(path)


def test_write_report_is_deterministic():
    """Same figures input must produce identical xlsx content."""
    from src.report.writer import write_report
    import openpyxl
    figures = [
        Figure(figure="allocation_sgs", value="35.0%", utilization="58.3%", status="OK",
               limit="20–60%", graph_path="(Position:SGS-01)->AssetClass",
               citation={"source_doc": "g.pdf", "page": 2, "chunk_id": "abc"}),
    ]
    paths = []
    for _ in range(2):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            paths.append(f.name)
    try:
        write_report(figures, paths[0])
        write_report(figures, paths[1])
        wb1 = openpyxl.load_workbook(paths[0])
        wb2 = openpyxl.load_workbook(paths[1])
        ws1 = wb1.active
        ws2 = wb2.active
        rows1 = list(ws1.iter_rows(values_only=True))
        rows2 = list(ws2.iter_rows(values_only=True))
        assert rows1 == rows2, "Non-deterministic output"
    finally:
        for p in paths:
            os.unlink(p)
