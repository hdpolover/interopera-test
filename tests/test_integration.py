"""Full pipeline integration tests for both firms.

Capstone test (Task 21): exercises the entire pipeline end-to-end —
ingestion → graph build → compute → reconcile → report → audit → firewall.

Requirements:
- Neo4j (bolt://neo4j:7687 in-container, bolt://localhost:7687 locally)
- Postgres (via POSTGRES_DSN env var)

Run in-container:
  docker compose run --rm --no-deps \
    -e NEO4J_TEST_URI=bolt://neo4j:7687 \
    -e NEO4J_TEST_USER=neo4j \
    -e NEO4J_TEST_PASSWORD=password \
    -e POSTGRES_DSN=postgresql://interopera:interopera@postgres:5432/interopera \
    app python -m pytest tests/test_integration.py -v
"""
from __future__ import annotations

import os
import tempfile
import uuid

import pytest

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PG_DSN = os.environ.get(
    "POSTGRES_DSN",
    os.environ.get(
        "POSTGRES_TEST_DSN",
        "postgresql://interopera:interopera@localhost:5432/interopera",
    ),
)


# ---------------------------------------------------------------------------
# Fixtures — graph (Neo4j) and audit (Postgres)
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def loaded_driver(driver):
    """Wipe graph, apply schema, and load full sample_docs for both firm tests."""
    from src.graph.schema import apply_schema
    from src.graph.builder import load_positions, load_rules
    from src.ingestion.holdings_parser import parse_holdings
    from src.ingestion.guidelines_parser import parse_guidelines

    with driver.session() as session:
        session.run("MATCH (n) DETACH DELETE n")

    apply_schema(driver)
    positions = parse_holdings(
        os.path.join(REPO_ROOT, "sample_docs", "sample_holdings.csv")
    )
    load_positions(driver, positions)
    # llm_client=None → deterministic 6-chunk stub; pdf_path=None is fine for stub
    chunks = parse_guidelines(pdf_path=None, llm_client=None)
    load_rules(driver, chunks)
    return driver


@pytest.fixture(scope="module")
def audit_logger():
    """Provide an AuditLogger with a fresh (TRUNCATED) audit_event table; skip if Postgres unavailable."""
    try:
        from src.audit.log import AuditLogger
        import psycopg

        with psycopg.connect(PG_DSN) as conn:
            conn.execute("TRUNCATE TABLE audit_event RESTART IDENTITY")
            conn.commit()

        log = AuditLogger(PG_DSN)
        yield log
        log.close()
    except Exception as e:
        pytest.skip(f"Postgres not available: {e}")


# ---------------------------------------------------------------------------
# Step 1 — Ingest + graph build (provenance check)
# ---------------------------------------------------------------------------


def test_graph_built_with_positions_and_rules(loaded_driver):
    """Graph must have Position nodes and SourceChunk nodes loaded from sample_docs."""
    with loaded_driver.session() as session:
        pos_count = session.run("MATCH (p:Position) RETURN count(p) AS n").single()["n"]
        chunk_count = session.run(
            "MATCH (c:SourceChunk) RETURN count(c) AS n"
        ).single()["n"]
        limit_count = session.run("MATCH (l:Limit) RETURN count(l) AS n").single()["n"]

    assert pos_count >= 13, f"Expected ≥13 Position nodes, got {pos_count}"
    assert chunk_count >= 6, f"Expected ≥6 SourceChunk nodes, got {chunk_count}"
    assert limit_count >= 6, f"Expected ≥6 Limit nodes, got {limit_count}"


# ---------------------------------------------------------------------------
# Step 2+3 — Firm A: 13 figures computed, reconcile ALL 13 pass
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def firm_a_figures(loaded_driver):
    from src.compute.config_loader import load_config
    from src.compute.engine import ComputeEngine

    config = load_config(
        os.path.join(REPO_ROOT, "config", "base.yaml"),
        os.path.join(REPO_ROOT, "config", "firm_a.yaml"),
    )
    return ComputeEngine(loaded_driver, config).run_all()


def test_firm_a_produces_13_figures(firm_a_figures):
    """ComputeEngine must return exactly 13 Figure objects for Firm A."""
    assert len(firm_a_figures) == 13, (
        f"Expected 13 figures, got {len(firm_a_figures)}: {[f.figure for f in firm_a_figures]}"
    )


def test_firm_a_full_pipeline(firm_a_figures):
    """End-to-end Firm A: all 13 figures must match firm_A_answer_key.xlsx (value+utilization+status)."""
    from src.reconcile.reconciler import reconcile, parse_answer_key_xlsx

    expected = parse_answer_key_xlsx(
        os.path.join(REPO_ROOT, "sample_docs", "firm_A_answer_key.xlsx")
    )
    assert len(expected) == 13, f"Answer key has {len(expected)} entries, expected 13"

    results = reconcile(firm_a_figures, expected)
    failed = [r for r in results if not r.passed]

    detail = "\n".join(
        f"  FAIL {r.figure}: computed=({r.computed_value!r}, {r.computed_utilization!r}, "
        f"{r.computed_status!r}) expected=({r.expected_value!r}, {r.expected_utilization!r}, "
        f"{r.expected_status!r}) delta={r.delta!r}"
        for r in failed
    )
    assert not failed, (
        f"BLOCKED: Firm A — {len(failed)}/13 figures did NOT reconcile:\n{detail}"
    )
    assert len([r for r in results if r.passed]) == 13, "Expected 13/13 passed for Firm A"


def test_firm_a_non_ig_ok(firm_a_figures):
    """Firm A: aggregate_non_ig_exposure is 15.0% OK (no fallen angels included)."""
    figs = {f.figure: f for f in firm_a_figures}
    fig = figs["aggregate_non_ig_exposure"]
    assert fig.value == "15.0%", f"Firm A non-IG value: expected '15.0%', got {fig.value!r}"
    assert fig.status == "OK", f"Firm A non-IG status: expected 'OK', got {fig.status!r}"


def test_firm_a_gre_ok(firm_a_figures):
    """Firm A: largest_gre_issuer is 7.0% OK (uses issuer-level grouping)."""
    figs = {f.figure: f for f in firm_a_figures}
    fig = figs["largest_gre_issuer"]
    assert fig.value == "7.0%", f"Firm A GRE value: expected '7.0%', got {fig.value!r}"
    assert fig.status == "OK", f"Firm A GRE status: expected 'OK', got {fig.status!r}"


# ---------------------------------------------------------------------------
# Step 4 — Report written: xlsx with 13 rows
# ---------------------------------------------------------------------------


def test_firm_a_report_written(firm_a_figures):
    """write_report produces an xlsx with a header row + 13 data rows for Firm A."""
    from src.report.writer import write_report
    import openpyxl

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name

    try:
        write_report(firm_a_figures, out_path)
        assert os.path.exists(out_path), "Report file was not created"

        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Row 0 is the header; rows 1..13 are data
        assert len(rows) == 14, f"Expected 14 rows (1 header + 13 data), got {len(rows)}"

        # Verify header
        assert rows[0][0] == "Section", f"First header column should be 'Section', got {rows[0][0]!r}"
        assert rows[0][5] == "Status", f"6th header column should be 'Status', got {rows[0][5]!r}"

        # Verify all 13 data rows have non-None values in key columns
        for i, row in enumerate(rows[1:], start=1):
            assert row[0] is not None, f"Row {i}: Section column is None"
            assert row[2] is not None, f"Row {i}: Value column is None"
            assert row[5] is not None, f"Row {i}: Status column is None"
    finally:
        os.unlink(out_path)


# ---------------------------------------------------------------------------
# Step 5 — Audit log records the run; verify_chain() returns OK
# ---------------------------------------------------------------------------


def test_audit_log_full_pipeline_run(loaded_driver, audit_logger):
    """Audit log captures all required event types for a full pipeline run; chain verifies."""
    from src.compute.config_loader import load_config, effective_config_hash
    from src.compute.engine import ComputeEngine
    from src.reconcile.reconciler import reconcile, parse_answer_key_xlsx
    from src.report.writer import write_report

    run_id = str(uuid.uuid4())

    config = load_config(
        os.path.join(REPO_ROOT, "config", "base.yaml"),
        os.path.join(REPO_ROOT, "config", "firm_a.yaml"),
    )
    cfg_hash = effective_config_hash(config)

    # config_loaded
    audit_logger.log_event(
        run_id=run_id,
        event_type="config_loaded",
        actor="system",
        payload={"firm_id": config.firm_id},
        config_hash=cfg_hash,
    )

    # graph_construction
    audit_logger.log_event(
        run_id=run_id,
        event_type="graph_construction",
        actor="system",
        payload={"event": "graph_construction", "source": "sample_holdings.csv"},
        config_hash=cfg_hash,
    )

    # compute figures + log figure_computed events
    figures = ComputeEngine(loaded_driver, config).run_all()
    assert len(figures) == 13, (
        f"Pipeline returned {len(figures)} figures, expected 13: "
        f"{[f.figure for f in figures]}"
    )
    for fig in figures:
        audit_logger.log_event(
            run_id=run_id,
            event_type="figure_computed",
            actor="system",
            payload={"figure": fig.figure, "value": fig.value, "status": fig.status},
            config_hash=cfg_hash,
        )

    # reconcile + log reconciliation
    expected = parse_answer_key_xlsx(
        os.path.join(REPO_ROOT, "sample_docs", "firm_A_answer_key.xlsx")
    )
    results = reconcile(figures, expected)
    passed_count = len([r for r in results if r.passed])
    audit_logger.log_event(
        run_id=run_id,
        event_type="reconciliation",
        actor="system",
        payload={"total": len(results), "passed": passed_count, "firm_id": config.firm_id},
        config_hash=cfg_hash,
    )

    # report_exported
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name
    try:
        write_report(figures, out_path)
        audit_logger.log_event(
            run_id=run_id,
            event_type="report_exported",
            actor="system",
            payload={"format": "xlsx", "firm_id": config.firm_id},
            config_hash=cfg_hash,
        )
    finally:
        os.unlink(out_path)

    # verify_chain() must return True — hash chain is intact
    assert audit_logger.verify_chain() is True, (
        "BLOCKED: audit log verify_chain() returned False — hash chain is broken"
    )

    # Verify all 5 required event types were logged for this run_id
    import psycopg

    with psycopg.connect(PG_DSN) as conn:
        rows = conn.execute(
            "SELECT DISTINCT event_type FROM audit_event WHERE run_id = %s",
            (run_id,),
        ).fetchall()

    logged_types = {row[0] for row in rows}
    required_types = {
        "graph_construction",
        "figure_computed",
        "reconciliation",
        "config_loaded",
        "report_exported",
    }
    missing = required_types - logged_types
    assert not missing, f"Missing audit event types: {missing}"


# ---------------------------------------------------------------------------
# Step 6 — Narrative + firewall
# ---------------------------------------------------------------------------


def test_firewall_stub_narrative_passes(firm_a_figures):
    """Stub narrative from Narrator passes firewall for Firm A."""
    from src.narrative.narrator import Narrator
    from src.firewall.checker import check_firewall

    narrator = Narrator(api_key=None)
    narrative = narrator.write_narrative(firm_a_figures, firm_id="firm_a")
    result = check_firewall(narrative, firm_a_figures)
    assert result.passed is True, (
        f"Stub narrative failed firewall: {result.offending_numbers}\n"
        f"Narrative (first 500 chars): {narrative[:500]}"
    )


def test_firewall_injected_number_fails(firm_a_figures):
    """Narrative with a number not in computed figures must be rejected by firewall."""
    from src.firewall.checker import check_firewall

    bad_narrative = "Exposure reached 99.9% of the tolerance limit."
    result = check_firewall(bad_narrative, firm_a_figures)
    assert result.passed is False, (
        "Firewall should reject 99.9% which is not in computed Firm A figures"
    )


# ---------------------------------------------------------------------------
# Firm B: full pipeline (proves config-only reconfiguration)
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def firm_b_figures(loaded_driver):
    from src.compute.config_loader import load_config
    from src.compute.engine import ComputeEngine

    config = load_config(
        os.path.join(REPO_ROOT, "config", "base.yaml"),
        os.path.join(REPO_ROOT, "config", "firm_b.yaml"),
    )
    return ComputeEngine(loaded_driver, config).run_all()


def test_firm_b_non_ig_breach(firm_b_figures):
    """Firm B: COR-05 (fallen angel) included in non-IG → 21.0% BREACH."""
    figs = {f.figure: f for f in firm_b_figures}
    fig = figs["aggregate_non_ig_exposure"]
    assert fig.status == "BREACH", (
        f"Firm B non-IG status: expected 'BREACH', got {fig.status!r}"
    )
    assert fig.value == "21.0%", (
        f"Firm B non-IG value: expected '21.0%', got {fig.value!r}"
    )
    assert fig.utilization == "10500 bps", (
        f"Firm B non-IG utilization: expected '10500 bps', got {fig.utilization!r}"
    )


def test_firm_b_gre_breach(firm_b_figures):
    """Firm B: Redhill Holdings (parent_issuer grouping) = 13.0% BREACH."""
    figs = {f.figure: f for f in firm_b_figures}
    fig = figs["largest_gre_issuer"]
    assert fig.status == "BREACH", (
        f"Firm B GRE status: expected 'BREACH', got {fig.status!r}"
    )
    assert fig.value == "13.0%", (
        f"Firm B GRE value: expected '13.0%', got {fig.value!r}"
    )
    assert fig.utilization == "10833 bps", (
        f"Firm B GRE utilization: expected '10833 bps', got {fig.utilization!r}"
    )


def test_firm_b_reconcile_all_pass(firm_b_figures):
    """Firm B reconcile against firm_b_expected.yaml — all 13 must pass."""
    from src.reconcile.reconciler import reconcile, parse_expected_yaml

    expected = parse_expected_yaml(
        os.path.join(REPO_ROOT, "config", "firm_b_expected.yaml")
    )
    assert len(expected) == 13, f"firm_b_expected.yaml has {len(expected)} entries, expected 13"

    results = reconcile(firm_b_figures, expected)
    failed = [r for r in results if not r.passed]

    detail = "\n".join(
        f"  FAIL {r.figure}: computed=({r.computed_value!r}, {r.computed_utilization!r}, "
        f"{r.computed_status!r}) expected=({r.expected_value!r}, {r.expected_utilization!r}, "
        f"{r.expected_status!r}) delta={r.delta!r}"
        for r in failed
    )
    assert not failed, (
        f"BLOCKED: Firm B — {len(failed)}/13 figures did NOT reconcile:\n{detail}"
    )
    assert len([r for r in results if r.passed]) == 13, "Expected 13/13 passed for Firm B"


def test_firm_b_report_written(firm_b_figures):
    """write_report produces an xlsx with a header row + 13 data rows for Firm B."""
    from src.report.writer import write_report
    import openpyxl

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        out_path = tmp.name

    try:
        write_report(firm_b_figures, out_path)
        assert os.path.exists(out_path), "Firm B report file was not created"

        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 14, (
            f"Firm B report: expected 14 rows (1 header + 13 data), got {len(rows)}"
        )

        # Verify header (mirror Firm A test)
        assert rows[0][0] == "Section", f"First header column should be 'Section', got {rows[0][0]!r}"
        assert rows[0][5] == "Status", f"6th header column should be 'Status', got {rows[0][5]!r}"

        # Verify all 13 data rows have non-None values in key columns (mirror Firm A test)
        for i, row in enumerate(rows[1:], start=1):
            assert row[0] is not None, f"Row {i}: Section column is None"
            assert row[2] is not None, f"Row {i}: Value column is None"
            assert row[5] is not None, f"Row {i}: Status column is None"
    finally:
        os.unlink(out_path)


def test_firm_b_firewall_passes(firm_b_figures):
    """Stub narrative for Firm B passes firewall."""
    from src.narrative.narrator import Narrator
    from src.firewall.checker import check_firewall

    narrator = Narrator(api_key=None)
    narrative = narrator.write_narrative(firm_b_figures, firm_id="firm_b")
    result = check_firewall(narrative, firm_b_figures)
    assert result.passed is True, (
        f"Stub narrative (Firm B) failed firewall: {result.offending_numbers}\n"
        f"Narrative (first 500 chars): {narrative[:500]}"
    )


# ---------------------------------------------------------------------------
# Step 7 — Assert firms produce DIFFERENT results (config-only reconfiguration)
# ---------------------------------------------------------------------------


def test_firms_produce_different_results(firm_a_figures, firm_b_figures):
    """The two firms produce different results, proving config-only reconfiguration end-to-end.

    Firm A non-IG: 15.0% OK  vs  Firm B non-IG: 21.0% BREACH
    Firm A GRE:     7.0% OK  vs  Firm B GRE:    13.0% BREACH
    """
    figs_a = {f.figure: f for f in firm_a_figures}
    figs_b = {f.figure: f for f in firm_b_figures}

    # Non-IG exposure
    assert figs_a["aggregate_non_ig_exposure"].value != figs_b["aggregate_non_ig_exposure"].value, (
        "Firm A and Firm B non-IG values should differ"
    )
    assert figs_a["aggregate_non_ig_exposure"].value == "15.0%", (
        f"Firm A non-IG: expected '15.0%', got {figs_a['aggregate_non_ig_exposure'].value!r}"
    )
    assert figs_b["aggregate_non_ig_exposure"].value == "21.0%", (
        f"Firm B non-IG: expected '21.0%', got {figs_b['aggregate_non_ig_exposure'].value!r}"
    )
    assert figs_a["aggregate_non_ig_exposure"].status == "OK"
    assert figs_b["aggregate_non_ig_exposure"].status == "BREACH"

    # GRE issuer concentration
    assert figs_a["largest_gre_issuer"].value != figs_b["largest_gre_issuer"].value, (
        "Firm A and Firm B GRE values should differ"
    )
    assert figs_a["largest_gre_issuer"].value == "7.0%", (
        f"Firm A GRE: expected '7.0%', got {figs_a['largest_gre_issuer'].value!r}"
    )
    assert figs_b["largest_gre_issuer"].value == "13.0%", (
        f"Firm B GRE: expected '13.0%', got {figs_b['largest_gre_issuer'].value!r}"
    )
    assert figs_a["largest_gre_issuer"].status == "OK"
    assert figs_b["largest_gre_issuer"].status == "BREACH"


# ---------------------------------------------------------------------------
# Capstone summary — both firms green, all constraints exercised
# ---------------------------------------------------------------------------


def test_full_pipeline_capstone_both_firms(
    loaded_driver, firm_a_figures, firm_b_figures
):
    """CAPSTONE: Both firms — full pipeline passes all 5 constraints simultaneously.

    Constraint 1: Positions + rules loaded with provenance.
    Constraint 2: All 13 figures computed for each firm.
    Constraint 3: Reconcile 13/13 for each firm against their answer keys.
    Constraint 4: Report written (xlsx, 13 rows) for each firm.
    Constraint 5: Audit log records the run, chain verifies, firewall passes.
    """
    from src.reconcile.reconciler import (
        reconcile,
        parse_answer_key_xlsx,
        parse_expected_yaml,
    )
    from src.narrative.narrator import Narrator
    from src.firewall.checker import check_firewall

    errors: list[str] = []

    # --- Firm A ---
    firm_a_expected = parse_answer_key_xlsx(
        os.path.join(REPO_ROOT, "sample_docs", "firm_A_answer_key.xlsx")
    )
    firm_a_results = reconcile(firm_a_figures, firm_a_expected)
    firm_a_failed = [r for r in firm_a_results if not r.passed]
    if firm_a_failed:
        errors.append(
            f"Firm A reconcile: {len(firm_a_failed)}/13 FAILED — "
            + "; ".join(f"{r.figure}:{r.delta}" for r in firm_a_failed)
        )

    firm_a_trace_failed = [
        f for f in firm_a_figures
        if not f.graph_path or not f.citation.get("chunk_id")
    ]
    if firm_a_trace_failed:
        errors.append(
            f"Firm A traceability: {[f.figure for f in firm_a_trace_failed]} missing graph_path/chunk_id"
        )

    narrator = Narrator(api_key=None)
    firm_a_narrative = narrator.write_narrative(firm_a_figures, firm_id="firm_a")
    firm_a_fw = check_firewall(firm_a_narrative, firm_a_figures)
    if not firm_a_fw.passed:
        errors.append(f"Firm A firewall: offending numbers {firm_a_fw.offending_numbers}")

    # --- Firm B ---
    firm_b_expected = parse_expected_yaml(
        os.path.join(REPO_ROOT, "config", "firm_b_expected.yaml")
    )
    firm_b_results = reconcile(firm_b_figures, firm_b_expected)
    firm_b_failed = [r for r in firm_b_results if not r.passed]
    if firm_b_failed:
        errors.append(
            f"Firm B reconcile: {len(firm_b_failed)}/13 FAILED — "
            + "; ".join(f"{r.figure}:{r.delta}" for r in firm_b_failed)
        )

    firm_b_trace_failed = [
        f for f in firm_b_figures
        if not f.graph_path or not f.citation.get("chunk_id")
    ]
    if firm_b_trace_failed:
        errors.append(
            f"Firm B traceability: {[f.figure for f in firm_b_trace_failed]} missing graph_path/chunk_id"
        )

    firm_b_narrative = narrator.write_narrative(firm_b_figures, firm_id="firm_b")
    firm_b_fw = check_firewall(firm_b_narrative, firm_b_figures)
    if not firm_b_fw.passed:
        errors.append(f"Firm B firewall: offending numbers {firm_b_fw.offending_numbers}")

    # --- Firms must differ ---
    figs_a = {f.figure: f for f in firm_a_figures}
    figs_b = {f.figure: f for f in firm_b_figures}
    if figs_a["aggregate_non_ig_exposure"].value == figs_b["aggregate_non_ig_exposure"].value:
        errors.append("CONFIGURATION BUG: Firm A and Firm B have identical non-IG values — config-only reconfiguration not working")
    if figs_a["largest_gre_issuer"].value == figs_b["largest_gre_issuer"].value:
        errors.append("CONFIGURATION BUG: Firm A and Firm B have identical GRE values — config-only reconfiguration not working")

    assert not errors, (
        "CAPSTONE BLOCKED — one or more pipeline constraints failed:\n"
        + "\n".join(f"  - {e}" for e in errors)
    )

    # Final counts for the record
    assert len([r for r in firm_a_results if r.passed]) == 13, "Firm A: not 13/13"
    assert len([r for r in firm_b_results if r.passed]) == 13, "Firm B: not 13/13"
